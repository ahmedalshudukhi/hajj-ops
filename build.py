#!/usr/bin/env python3
"""
Hajj Ops Dashboard — Data Builder v8 (v11.8 schema)

Three-type unit model: Leadership (4) + Command + Operational.
Operational split: Mike Medical (20) + Alpha Ambulance (48) + Romeo Foot-runner (49).
10 fixed sites: 9 clinical + OCC. 3 floating functions (Logistics, Training, Mobile).
v11.7: SRCA dropped (10 sites). Supervisors 8 → 6. Romeo 46 → 47 + Romeo-48 (single-para). Units 138 → 137.
v11.8: Romeo-48 size 1 → 2 (full unit). Romeo Solo-49 added (single-para, MIN3). Romeo 48 → 49. Units 137 → 138. Roster 275 → 277. Mandate paramedics = 250 exactly.
OCC accommodation: housed at the closest available HMG sites (primarily ARF3 and MUZ3), pending SAR confirmation of rooms in the ABG building.
"""
import os, sys, json, urllib.request, tempfile
from datetime import datetime, timezone, time as dtime
from collections import defaultdict, Counter
import openpyxl

# Source: Backend Google Sheet (live edits) — export as xlsx on every build.
# This replaces the older static xlsx file (1USar5...) which would go stale
# whenever Ahmed edited the Backend Sheet without re-exporting.
DEFAULT_FILE_ID = "16nlZuencav9uB9o9Kscgmb5UvVGeKcu4e3YxqdKohiw"
FILE_ID = os.environ.get("GDRIVE_FILE_ID", DEFAULT_FILE_ID)
DOWNLOAD_URL = f"https://docs.google.com/spreadsheets/d/{FILE_ID}/export?format=xlsx"

STATIONS = ["ARF1","ARF2","ARF3","MUZ1","MUZ2","MUZ3","MIN1","MIN2","MIN3"]
SITES_ALL = STATIONS + ["OCC"]                    # 10 fixed sites (SRCA dropped v11.7)
SITES_FLOATING = ["Logistics","Training","Mobile"]  # 3 floating functions (no fixed home)

# ─── Static project data ──────────────────────────────────────────
CALENDAR = [
    {"dh":1,"greg":"Mon 18 May 2026","day":"Mon","phase":"Site Prep","events":"Sites ready"},
    {"dh":2,"greg":"Tue 19 May 2026","day":"Tue","phase":"Site Prep","events":""},
    {"dh":3,"greg":"Wed 20 May 2026","day":"Wed","phase":"Site Prep","events":""},
    {"dh":4,"greg":"Thu 21 May 2026","day":"Thu","phase":"Training","events":"Staff mobilization"},
    {"dh":5,"greg":"Fri 22 May 2026","day":"Fri","phase":"Training","events":""},
    {"dh":6,"greg":"Sat 23 May 2026","day":"Sat","phase":"Training","events":""},
    {"dh":7,"greg":"Sun 24 May 2026","day":"Sun","phase":"Training","events":"Drills"},
    {"dh":8,"greg":"Mon 25 May 2026","day":"Mon","phase":"Ops Day 1","events":"Tarwiyah · B1A 20:00"},
    {"dh":9,"greg":"Tue 26 May 2026","day":"Tue","phase":"Ops Day 2","events":"ARAFAT · C CRITICAL"},
    {"dh":10,"greg":"Wed 27 May 2026","day":"Wed","phase":"Ops Day 3","events":"EID · D→E"},
    {"dh":11,"greg":"Thu 28 May 2026","day":"Thu","phase":"Ops Day 4","events":"E"},
    {"dh":12,"greg":"Fri 29 May 2026","day":"Fri","phase":"Ops Day 5","events":"E"},
    {"dh":13,"greg":"Sat 30 May 2026","day":"Sat","phase":"Ops Day 6","events":"E ends 18:00"},
    {"dh":14,"greg":"Sun 31 May 2026","day":"Sun","phase":"Demob","events":"Demob"},
    {"dh":15,"greg":"Mon 1 Jun 2026","day":"Mon","phase":"Demob","events":"Demob"},
]
TIMELINE = [
    {"phase":"0","activity":"Master plan to SAR","date_dh":"~10 DQ","greg":"~27 Apr","owner":"PM"},
    {"phase":"0","activity":"Equipment procured","date_dh":"~20 DQ","greg":"~7 May","owner":"LOG"},
    {"phase":"1","activity":"SAR handover","date_dh":"~25 DQ","greg":"~12 May","owner":"PM"},
    {"phase":"1","activity":"Sites ready","date_dh":"1 DH","greg":"18 May","owner":"LOG"},
    {"phase":"2","activity":"Staff mobilization","date_dh":"4 DH","greg":"21 May","owner":"LOG"},
    {"phase":"2","activity":"Drills","date_dh":"5-6 DH","greg":"22-23 May","owner":"TRN"},
    {"phase":"3","activity":"Ops B-E","date_dh":"8-13 DH","greg":"25-30 May","owner":"PM"},
    {"phase":"4","activity":"Demob","date_dh":"14-27 DH","greg":"31 May-13 Jun","owner":"LOG"},
    {"phase":"5","activity":"Final report","date_dh":"30 days post","greg":"","owner":"PM"},
]
MOVEMENTS = [
    {"code":"PRE-B","shift":"DAY","dh":"4-7 DH","label":"Pre-Boarding setup","staff":248,"para":248,"amb":12},
    {"code":"B1A","shift":"NIGHT","dh":"8/9 DH 20:00-01:00","label":"Boarding 1A","staff":248,"para":248,"amb":18},
    {"code":"B1B","shift":"NIGHT","dh":"9 DH 02:00-04:00","label":"Boarding 1B","staff":248,"para":248,"amb":18},
    {"code":"B2A","shift":"DAY","dh":"9 DH 05:00-07:00","label":"Boarding 2A","staff":248,"para":248,"amb":18},
    {"code":"B2B","shift":"DAY","dh":"9 DH 08:00-10:00","label":"Boarding 2B","staff":248,"para":248,"amb":18},
    {"code":"GAP","shift":"DAY","dh":"9 DH 13:00-19:00","label":"Inter-movement Gap","staff":248,"para":248,"amb":25},
    {"code":"C","shift":"NIGHT","dh":"9/10 DH","label":"Day of Arafat / Nafra","staff":248,"para":248,"amb":25},
    {"code":"D","shift":"NIGHT","dh":"10 DH 00:00-08:00","label":"Muz → Mina","staff":248,"para":248,"amb":25},
    {"code":"E1","shift":"DAY","dh":"10 DH 09:00-12:00","label":"Mina ops — E1 evac","staff":248,"para":248,"amb":25},
    {"code":"E2","shift":"DAY","dh":"10 DH 13:00-17:00","label":"Mina ops — E2","staff":160,"para":160,"amb":18},
    {"code":"E3","shift":"MIXED","dh":"11-13 DH","label":"Mina ops — E3 sustain","staff":160,"para":160,"amb":18},
    {"code":"DEMOB","shift":"DAY","dh":"14+ DH","label":"Demobilization","staff":160,"para":160,"amb":12},
]
MOVEMENT_PHASES = [
    # ─── Metro movements (per SAR official ops doc, 1447H) ──────────
    # Movement A — regular metro, all 9 stations, both platforms
    # (operators and tawafa shuttle)
    {"mvt":"A",  "start_dh":"7 DH",  "start_hour":"08:00", "end_dh":"8 DH",  "end_hour":"00:00", "shift":"MIXED", "trains":7,  "desc":"Regular metro — all stations"},
    {"mvt":"A",  "start_dh":"8 DH",  "start_hour":"00:00", "end_dh":"8 DH",  "end_hour":"02:00", "shift":"NIGHT", "trains":7,  "desc":"Regular metro — all stations"},
    {"mvt":"A",  "start_dh":"8 DH",  "start_hour":"04:00", "end_dh":"8 DH",  "end_hour":"16:00", "shift":"DAY",   "trains":7,  "desc":"Regular metro — all stations"},

    # Movement B — Convoy: Mina (south side) → Arafat (north/south)
    # Four sub-phases B1A → B1B → B2A → B2B as pilgrims ascend.
    {"mvt":"B1A","start_dh":"8 DH",  "start_hour":"18:00", "end_dh":"9 DH",  "end_hour":"02:00", "shift":"NIGHT", "trains":12, "desc":"Convoy ascent — Mina(S) → Arafat(N), via Muz3"},
    {"mvt":"B1B","start_dh":"9 DH",  "start_hour":"02:00", "end_dh":"9 DH",  "end_hour":"05:00", "shift":"NIGHT", "trains":15, "desc":"Convoy ascent — Mina(S) → Arafat(N), incl. Mina3"},
    {"mvt":"B2A","start_dh":"9 DH",  "start_hour":"05:30", "end_dh":"9 DH",  "end_hour":"08:00", "shift":"DAY",   "trains":15, "desc":"Convoy ascent — Mina(S) → Arafat(S+N), via Muz3"},
    {"mvt":"B2B","start_dh":"9 DH",  "start_hour":"08:00", "end_dh":"9 DH",  "end_hour":"11:00", "shift":"DAY",   "trains":15, "desc":"Convoy ascent — Mina(S) → Arafat(S only)"},

    # Movement C — Shuttle convoy / Nafra: Arafat → Muzdalifah
    # Critical 5h window after sunset on Day of Arafah.
    {"mvt":"C",  "start_dh":"9 DH",  "start_hour":"18:57", "end_dh":"10 DH", "end_hour":"00:30", "shift":"NIGHT", "trains":12, "desc":"Nafra shuttle — Arafat → Muzdalifah (3 paired stations)"},

    # Movement D — Skip-stop: Muzdalifah → Jamarat
    {"mvt":"D",  "start_dh":"10 DH", "start_hour":"01:00", "end_dh":"10 DH", "end_hour":"09:00", "shift":"NIGHT", "trains":12, "desc":"Skip-stop — Muzdalifah → Jamarat (3 patterns)"},

    # Movement E — Jamarat shuttling (DH 10-13), with daily 02:00-04:00 maintenance
    {"mvt":"E1", "start_dh":"10 DH", "start_hour":"09:00", "end_dh":"11 DH", "end_hour":"02:00", "shift":"MIXED", "trains":12, "desc":"Metro-type — Jamarat throwing (E1: 5 stations)"},
    {"mvt":"E1", "start_dh":"11 DH", "start_hour":"04:00", "end_dh":"11 DH", "end_hour":"11:00", "shift":"DAY",   "trains":12, "desc":"Metro-type — Jamarat throwing"},
    {"mvt":"E2", "start_dh":"11 DH", "start_hour":"11:00", "end_dh":"11 DH", "end_hour":"17:00", "shift":"DAY",   "trains":12, "desc":"Metro-type — Jamarat (E2: dedicated north trains)"},
    {"mvt":"E1", "start_dh":"11 DH", "start_hour":"17:00", "end_dh":"12 DH", "end_hour":"02:00", "shift":"MIXED", "trains":12, "desc":"Metro-type — Jamarat throwing"},
    {"mvt":"E1", "start_dh":"12 DH", "start_hour":"04:00", "end_dh":"12 DH", "end_hour":"11:00", "shift":"DAY",   "trains":12, "desc":"Metro-type — Jamarat throwing"},
    {"mvt":"E2", "start_dh":"12 DH", "start_hour":"11:00", "end_dh":"12 DH", "end_hour":"17:00", "shift":"DAY",   "trains":12, "desc":"Metro-type — Jamarat (E2: dedicated north trains)"},
    {"mvt":"E1", "start_dh":"12 DH", "start_hour":"17:00", "end_dh":"13 DH", "end_hour":"02:00", "shift":"MIXED", "trains":12, "desc":"Metro-type — Jamarat throwing"},
    {"mvt":"E1", "start_dh":"13 DH", "start_hour":"04:00", "end_dh":"13 DH", "end_hour":"18:00", "shift":"MIXED", "trains":12, "desc":"Metro-type — Jamarat throwing (final)"},

    # ─── Medical operational phases (fallback labels for non-train hours) ──
    # These appear AFTER metro entries so the metro label wins when both
    # match (e.g. hour DH 7 12:00 → "A" not "PRE-B").
    {"mvt":"PRE-B","start_dh":"4 DH","start_hour":"06:00", "end_dh":"8 DH",  "end_hour":"18:00", "shift":"DAY",   "trains":0,  "desc":"Pre-mobilization (medical setup)"},
    {"mvt":"DEMOB","start_dh":"14 DH","start_hour":"00:00","end_dh":"14 DH", "end_hour":"23:00", "shift":"DAY",   "trains":0,  "desc":"Demobilization"},
]

# ─── SAR Metro reference data (from official 1447H operations doc) ───────
# Per-movement, per-station platform activity. Platform codes:
#   'N' = northern platform (الرصيف الشمالي)
#   'S' = southern platform (الرصيف الجنوبي)
#   'NS' = both platforms in use
# Role: 'board' (alighting passengers depart here), 'alight' (arrivals),
# 'pass' (train passes through but doesn't stop), 'both' (board + alight)
METRO_PLATFORMS = {
    "A": [  # Regular metro — all 9 stations, both platforms
        {"st":"ARF1","plat":"NS","role":"both"}, {"st":"ARF2","plat":"NS","role":"both"}, {"st":"ARF3","plat":"NS","role":"both"},
        {"st":"MUZ1","plat":"NS","role":"both"}, {"st":"MUZ2","plat":"NS","role":"both"}, {"st":"MUZ3","plat":"NS","role":"both"},
        {"st":"MIN1","plat":"NS","role":"both"}, {"st":"MIN2","plat":"NS","role":"both"}, {"st":"MIN3","plat":"NS","role":"both"},
    ],
    "B1A": [  # Mina(S) → Muz3 → Arafat(N) — convoy ascent, no Mina 3
        {"st":"MIN1","plat":"S","role":"board"}, {"st":"MIN2","plat":"S","role":"board"},
        {"st":"MUZ3","plat":"S","role":"board"},
        {"st":"ARF1","plat":"N","role":"alight"}, {"st":"ARF2","plat":"N","role":"alight"}, {"st":"ARF3","plat":"N","role":"alight"},
    ],
    "B1B": [  # Adds Mina 3; Arafat 3 also gets south side for Mina-3 passengers
        {"st":"MIN1","plat":"S","role":"board"}, {"st":"MIN2","plat":"S","role":"board"}, {"st":"MIN3","plat":"S","role":"board"},
        {"st":"MUZ3","plat":"S","role":"board"},
        {"st":"ARF1","plat":"N","role":"alight"}, {"st":"ARF2","plat":"N","role":"alight"}, {"st":"ARF3","plat":"NS","role":"alight"},
    ],
    "B2A": [  # Arrivals switch to south side at Arafat (north remains for Muz-3 traffic)
        {"st":"MIN1","plat":"S","role":"board"}, {"st":"MIN2","plat":"S","role":"board"}, {"st":"MIN3","plat":"S","role":"board"},
        {"st":"MUZ3","plat":"S","role":"board"},
        {"st":"ARF1","plat":"S","role":"alight"}, {"st":"ARF2","plat":"S","role":"alight"}, {"st":"ARF3","plat":"NS","role":"alight"},
    ],
    "B2B": [  # Mina only, all Arafat south
        {"st":"MIN1","plat":"S","role":"board"}, {"st":"MIN2","plat":"S","role":"board"}, {"st":"MIN3","plat":"S","role":"board"},
        {"st":"ARF1","plat":"S","role":"alight"}, {"st":"ARF2","plat":"S","role":"alight"}, {"st":"ARF3","plat":"S","role":"alight"},
    ],
    "C": [  # Nafra — paired stations Arafat-N → Muz-N
        {"st":"ARF1","plat":"NS","role":"board"}, {"st":"ARF2","plat":"NS","role":"board"}, {"st":"ARF3","plat":"NS","role":"board"},
        {"st":"MUZ1","plat":"NS","role":"alight"}, {"st":"MUZ2","plat":"NS","role":"alight"}, {"st":"MUZ3","plat":"NS","role":"alight"},
    ],
    "D": [  # Skip-stop: 3 patterns from each Muz to Jamarat with intermediate stops
        {"st":"MUZ1","plat":"N","role":"board"}, {"st":"MUZ2","plat":"N","role":"board"}, {"st":"MUZ3","plat":"N","role":"board"},
        {"st":"MIN1","plat":"N","role":"alight"}, {"st":"MIN2","plat":"N","role":"alight"}, {"st":"MIN3","plat":"N","role":"alight"},
    ],
    "E1": [  # Metro-type Jamarat: 5 stops only — Arafat 3, Muz 3, Mina 1, Mina 2, Jamarat
        {"st":"ARF3","plat":"NS","role":"both"}, {"st":"MUZ3","plat":"NS","role":"both"},
        {"st":"MIN1","plat":"NS","role":"both"}, {"st":"MIN2","plat":"NS","role":"both"}, {"st":"MIN3","plat":"NS","role":"both"},
    ],
    "E2": [  # Dedicated north-platform trains, each Arafat/Muz station has its own train to Jamarat
        {"st":"ARF1","plat":"N","role":"board"}, {"st":"ARF2","plat":"N","role":"board"}, {"st":"ARF3","plat":"N","role":"board"},
        {"st":"MUZ1","plat":"N","role":"board"}, {"st":"MUZ2","plat":"N","role":"board"}, {"st":"MUZ3","plat":"N","role":"board"},
        {"st":"MIN1","plat":"N","role":"both"}, {"st":"MIN2","plat":"N","role":"both"}, {"st":"MIN3","plat":"N","role":"alight"},
    ],
    "PRE-B": [],   # No train operations during medical pre-mob
    "DEMOB": [],   # No train operations during medical demob
}

# Hourly passenger flow per station (counts ARRIVING at the station during
# the movement, per SAR doc Section 4). Keyed by (dh, hour, station).
# Values in pilgrim-headcount. Used for the operational forecast view.
METRO_PAX_FLOW = {
    # Movement B — Mina → Arafat (counts of pilgrims arriving at the station)
    # Source: SAR doc page "تدفقات الحشود — الحركة B — أيام التشغيل 8 و 9"
    "B": [
        # (dh, hour, station, count)
        (8,18,"MUZ3",4630),
        (8,19,"MUZ3",3197),
        (8,20,"MUZ3",3001), (8,20,"MIN1",10694), (8,20,"MIN2",13412), (8,20,"MIN3",9847),
        (8,21,"MUZ3",1530), (8,21,"MIN1",9586),  (8,21,"MIN2",11922), (8,21,"MIN3",7727),
        (8,22,"MUZ3",2059), (8,22,"MIN1",8294),  (8,22,"MIN2",10565), (8,22,"MIN3",6972),
        (8,23,"MUZ3",1937), (8,23,"MIN1",7207),  (8,23,"MIN2",9024),  (8,23,"MIN3",3962),
        (9,0,"MIN1",6151),  (9,0,"MIN2",7708),   (9,0,"MIN3",5573),
        (9,2,"MUZ3",3589),  (9,2,"MIN1",4888),   (9,2,"MIN2",6768),   (9,2,"MIN3",5644),
        (9,3,"MUZ3",1942),  (9,3,"MIN1",2102),   (9,3,"MIN2",2068),   (9,3,"MIN3",5728),
        (9,4,"MUZ3",2508),  (9,4,"MIN3",4891),
        (9,5,"MUZ3",2383),  (9,5,"MIN1",7149),   (9,5,"MIN2",7314),   (9,5,"MIN3",1081),
        (9,6,"MUZ3",1977),  (9,6,"MIN1",13562),  (9,6,"MIN2",13006),
        (9,7,"MIN1",11521), (9,7,"MIN2",11719),
        (9,8,"MIN1",11061), (9,8,"MIN2",11591),
        (9,9,"MIN1",11144), (9,9,"MIN2",10992),
        (9,10,"MIN1",2068), (9,10,"MIN2",2256),
    ],
    # Movement C — Nafra (Arafat → Muzdalifah). Boarding counts at Arafat
    # stations during 5 tafweej periods (zones).
    "C": [
        # Period 1: 18:10-19:10 — N=Northern, S=Southern platform
        (9,18,"ARF1_N",13515),(9,18,"ARF1_S",12302),(9,18,"ARF2_N",14397),(9,18,"ARF2_S",15170),(9,18,"ARF3_N",10585),(9,18,"ARF3_S",12548),
        # Period 2: 19:30-20:10
        (9,19,"ARF1_N",11536),(9,19,"ARF1_S",12558),(9,19,"ARF2_N",12875),(9,19,"ARF2_S",10029),(9,19,"ARF3_N",9718), (9,19,"ARF3_S",11048),
        # Period 3: 20:30-21:20
        (9,20,"ARF1_N",13066),(9,20,"ARF1_S",12142),(9,20,"ARF2_N",12461),(9,20,"ARF2_S",12491),(9,20,"ARF3_N",10696),(9,20,"ARF3_S",11691),
        # Period 4: 21:30-22:20
        (9,21,"ARF1_N",14030),(9,21,"ARF1_S",16278),(9,21,"ARF2_N",10998),(9,21,"ARF2_S",11188),(9,21,"ARF3_N",7518), (9,21,"ARF3_S",6374),
        # Period 5: 22:30-23:20 (only ARF2 active)
        (9,22,"ARF2_N",10736),(9,22,"ARF2_S",8000),
    ],
    # Movement D — Muzdalifah → Jamarat. Pax arriving at each Muz station.
    "D": [
        (10,1,"MUZ1",6367),  (10,1,"MUZ2",12260), (10,1,"MUZ3",11630),
        (10,2,"MUZ1",2846),  (10,2,"MUZ2",12472), (10,2,"MUZ3",9551),
        (10,3,"MUZ1",2406),  (10,3,"MUZ2",12229), (10,3,"MUZ3",4820),
        (10,4,"MUZ1",2274),  (10,4,"MUZ2",12381), (10,4,"MUZ3",4931),
        (10,5,"MUZ1",19246), (10,5,"MUZ2",15482), (10,5,"MUZ3",5207),
        (10,6,"MUZ1",23073), (10,6,"MUZ2",16811), (10,6,"MUZ3",4591),
        (10,7,"MUZ1",22919), (10,7,"MUZ2",17019), (10,7,"MUZ3",4712),
        (10,8,"MUZ1",22468), (10,8,"MUZ2",16872), (10,8,"MUZ3",5131),
        (10,9,"MUZ1",3828),  (10,9,"MUZ2",2819),  (10,9,"MUZ3",852),
    ],
}

# Tafweej (boarding pass) zones for Arafat→Muzdalifah Nafra (Movement C).
# Five colored zones, each releases pilgrims in a defined window.
METRO_TAFWEEJ = [
    {"zone":1, "color":"#9ED65E", "color_name":"green",  "start":"18:10","end":"19:30","total":80166},
    {"zone":2, "color":"#FEFE19", "color_name":"yellow", "start":"19:10","end":"20:30","total":72412},
    {"zone":3, "color":"#1AB8EF", "color_name":"blue",   "start":"20:00","end":"21:30","total":73718},
    {"zone":4, "color":"#DC7BCF", "color_name":"purple", "start":"21:00","end":"22:30","total":57309},
    {"zone":5, "color":"#FFC518", "color_name":"orange", "start":"22:00","end":"23:30","total":32726},
]

# Station opening / closing times per movement (from SAR doc Section 2).
# Used to flag whether each station is "open" or "closed" during the
# selected hour. Times in 24h ISO HH:MM format, anchored to Hijri DH day.
METRO_STATION_HOURS = {
    "ARF1": [
        {"mvt":"A","dh":7,"open":"07:30","close":None},
        {"mvt":"A","dh":8,"open":None,"close":"01:30"}, {"mvt":"A","dh":8,"open":"03:30","close":None}, {"mvt":"A","dh":8,"open":None,"close":"15:30"},
        {"mvt":"B","dh":8,"open":"17:30","close":None},
        {"mvt":"B","dh":9,"open":None,"close":"11:45"},
        {"mvt":"C","dh":9,"open":"17:58","close":None},
        {"mvt":"C","dh":9,"open":None,"close":"23:30"},
    ],
    "ARF2": [{"mvt":"A","dh":7,"open":"07:30","close":None}, {"mvt":"A","dh":8,"open":None,"close":"01:30"}, {"mvt":"A","dh":8,"open":"03:30","close":None}, {"mvt":"A","dh":8,"open":None,"close":"15:30"}, {"mvt":"B","dh":8,"open":"17:30","close":None}, {"mvt":"B","dh":9,"open":None,"close":"11:45"}, {"mvt":"C","dh":9,"open":"17:58","close":None}, {"mvt":"C","dh":9,"open":None,"close":"23:30"}],
    "ARF3": [{"mvt":"A","dh":7,"open":"07:30","close":None}, {"mvt":"A","dh":8,"open":None,"close":"01:30"}, {"mvt":"A","dh":8,"open":"03:30","close":None}, {"mvt":"A","dh":8,"open":None,"close":"15:30"}, {"mvt":"B","dh":8,"open":"17:30","close":None}, {"mvt":"B","dh":9,"open":None,"close":"11:45"}, {"mvt":"C","dh":9,"open":"17:58","close":None}, {"mvt":"C","dh":9,"open":None,"close":"23:30"}],
    "MUZ1": [{"mvt":"A","dh":7,"open":"07:30","close":None}, {"mvt":"A","dh":8,"open":None,"close":"01:30"}, {"mvt":"A","dh":8,"open":"03:30","close":None}, {"mvt":"A","dh":8,"open":None,"close":"15:30"}, {"mvt":"C","dh":9,"open":"17:58","close":None}, {"mvt":"D","dh":10,"open":"00:00","close":None}, {"mvt":"D","dh":10,"open":None,"close":"08:30"}],
    "MUZ2": [{"mvt":"A","dh":7,"open":"07:30","close":None}, {"mvt":"A","dh":8,"open":None,"close":"01:30"}, {"mvt":"A","dh":8,"open":"03:30","close":None}, {"mvt":"A","dh":8,"open":None,"close":"15:30"}, {"mvt":"C","dh":9,"open":"17:58","close":None}, {"mvt":"D","dh":10,"open":"00:00","close":None}, {"mvt":"D","dh":10,"open":None,"close":"08:30"}],
    "MUZ3": [{"mvt":"A","dh":7,"open":"07:30","close":None}, {"mvt":"A","dh":8,"open":None,"close":"01:30"}, {"mvt":"A","dh":8,"open":"03:30","close":None}, {"mvt":"A","dh":8,"open":None,"close":"15:30"}, {"mvt":"B","dh":8,"open":"17:30","close":None}, {"mvt":"B","dh":9,"open":None,"close":"10:30"}, {"mvt":"C","dh":9,"open":"17:58","close":None}, {"mvt":"D","dh":10,"open":"00:00","close":None}, {"mvt":"D","dh":10,"open":None,"close":"08:30"}],
    "MIN1": [{"mvt":"A","dh":7,"open":"07:30","close":None}, {"mvt":"A","dh":8,"open":None,"close":"01:30"}, {"mvt":"A","dh":8,"open":"03:30","close":None}, {"mvt":"A","dh":8,"open":None,"close":"15:30"}, {"mvt":"B","dh":8,"open":"17:30","close":None}, {"mvt":"B","dh":9,"open":None,"close":"07:15"}, {"mvt":"D","dh":10,"open":"08:30","close":None}, {"mvt":"E","dh":10,"open":None,"close":None}],
    "MIN2": [{"mvt":"A","dh":7,"open":"07:30","close":None}, {"mvt":"A","dh":8,"open":None,"close":"01:30"}, {"mvt":"A","dh":8,"open":"03:30","close":None}, {"mvt":"A","dh":8,"open":None,"close":"15:30"}, {"mvt":"B","dh":8,"open":"17:30","close":None}, {"mvt":"B","dh":9,"open":None,"close":"10:30"}, {"mvt":"D","dh":10,"open":"08:30","close":None}, {"mvt":"E","dh":10,"open":None,"close":None}],
    "MIN3": [{"mvt":"A","dh":7,"open":"07:30","close":None}, {"mvt":"A","dh":8,"open":None,"close":"01:30"}, {"mvt":"A","dh":8,"open":"03:30","close":None}, {"mvt":"A","dh":8,"open":None,"close":"15:30"}, {"mvt":"B","dh":9,"open":None,"close":"01:30"}, {"mvt":"D","dh":10,"open":"08:30","close":None}, {"mvt":"E","dh":10,"open":None,"close":None}],
}
GP_COVERAGE = [
    {"station":"ARF1","gp_day":"GP-01","gp_night":"GP-11","covers":"Mike-1, Mike-2","notes":""},
    {"station":"ARF2","gp_day":"GP-02","gp_night":"GP-12","covers":"Mike-3, Mike-4","notes":""},
    {"station":"ARF3","gp_day":"GP-03","gp_night":"GP-13","covers":"Mike-5, Mike-6","notes":""},
    {"station":"MUZ1","gp_day":"GP-04","gp_night":"GP-14","covers":"Mike-7, Mike-8","notes":""},
    {"station":"MUZ2","gp_day":"GP-05","gp_night":"GP-15","covers":"Mike-9, Mike-10","notes":""},
    {"station":"MUZ3","gp_day":"GP-06","gp_night":"GP-16","covers":"Mike-11, Mike-12","notes":""},
    {"station":"MIN1","gp_day":"GP-07","gp_night":"GP-17","covers":"Mike-13, Mike-14","notes":""},
    {"station":"MIN2","gp_day":"GP-08","gp_night":"GP-18","covers":"Mike-15, Mike-16","notes":""},
    {"station":"MIN3","gp_day":"GP-09","gp_night":"GP-19","covers":"Mike-17, Mike-18 (Jamarat)","notes":""},
    {"station":"OCC","gp_day":"GP-10","gp_night":"GP-20","covers":"Mike-19, Mike-20 (OCC ops)","notes":""},
]
ACCOMMODATION = [
    {"location":"ARF1","sta_para":12,"amb_crew":8,"rov_fwd":2,"gps":2,"support":2,"total_beds":26,"bunk_sets":13},
    {"location":"ARF2","sta_para":12,"amb_crew":8,"rov_fwd":2,"gps":2,"support":1,"total_beds":25,"bunk_sets":13},
    {"location":"ARF3","sta_para":12,"amb_crew":8,"rov_fwd":2,"gps":2,"support":1,"total_beds":25,"bunk_sets":13},
    {"location":"MUZ1","sta_para":12,"amb_crew":8,"rov_fwd":2,"gps":2,"support":2,"total_beds":26,"bunk_sets":13},
    {"location":"MUZ2","sta_para":12,"amb_crew":8,"rov_fwd":2,"gps":2,"support":1,"total_beds":25,"bunk_sets":13},
    {"location":"MUZ3","sta_para":12,"amb_crew":8,"rov_fwd":2,"gps":2,"support":1,"total_beds":25,"bunk_sets":13},
    {"location":"MIN1","sta_para":12,"amb_crew":8,"rov_fwd":2,"gps":2,"support":2,"total_beds":26,"bunk_sets":13},
    {"location":"MIN2","sta_para":12,"amb_crew":8,"rov_fwd":2,"gps":2,"support":1,"total_beds":25,"bunk_sets":13},
    {"location":"MIN3","sta_para":12,"amb_crew":8,"rov_fwd":2,"gps":2,"support":1,"total_beds":25,"bunk_sets":13},
    {"location":"OCC","sta_para":4,"amb_crew":0,"rov_fwd":0,"gps":2,"support":4,"total_beds":10,"bunk_sets":5},
]

# ─── Helpers ──────────────────────────────────────────────────────
def num(v, default=0):
    if v is None or v == "": return default
    try: return float(v)
    except (TypeError, ValueError): return default

def s(v): return str(v).strip() if v is not None else ""

def compute_shift_duration(start_time, end_time):
    """Calculate shift duration in hours from start/end time cells.
    Handles midnight wrap (e.g. 19:00 → 07:00 = 12 h, 20:00 → 02:00 = 6 h).
    Treats same start==end as 24h (full-day rotation, e.g. 24/7 shift).
    Falls back to 0 for unparseable inputs.

    Why this exists: the 'Duration (h)' cells in the Shifts sheet are
    formatted as Excel time/date values (e.g. cells show as 12:00 but
    serialize as datetime.datetime(1900,1,12,0,0)), so the previous
    `num(r.get('Duration (h)'))` call returned 0 for nearly every shift.
    Always compute from Start/End instead — single source of truth.
    """
    if not hasattr(start_time, 'hour') or not hasattr(end_time, 'hour'):
        return 0.0
    sh_ = start_time.hour + start_time.minute / 60
    eh_ = end_time.hour + end_time.minute / 60
    if sh_ == eh_:
        return 24.0
    if eh_ < sh_:
        return (24 - sh_) + eh_
    return eh_ - sh_

def compute_phase_duration(start_dh, start_hour, end_dh, end_hour):
    """Calculate movement phase duration in hours (wall-clock elapsed) from
    start/end DH+hour strings.
    Args:
        start_dh / end_dh: '4 DH' format
        start_hour / end_hour: '06:00' format (24-hour)

    Why this exists: the hand-maintained `duration_hrs` field on each
    MOVEMENT_PHASES entry drifted from the start/end times (e.g. B1A was
    stored as 6 h but 8 DH 20:00 → 9 DH 01:00 is 5 h; PRE-B was stored as
    60 h but 4 DH 06:00 → 8 DH 17:00 is 107 h). Always derive at build time
    from start_dh/start_hour/end_dh/end_hour — single source of truth.
    """
    def _dh(x): return int(str(x).split()[0])
    def _hr(x):
        parts = str(x).split(':')
        return int(parts[0]) + (int(parts[1]) / 60 if len(parts) >= 2 else 0)
    return (_dh(end_dh) - _dh(start_dh)) * 24 + (_hr(end_hour) - _hr(start_hour))

def download_xlsx():
    # Order of preference:
    #   1. Local backend.xlsx / _backend_cache.xlsx (Ahmed dropped a fresh export)
    #   2. Service-account-authenticated Drive export (cron-friendly, secure)
    #   3. Public /export?format=xlsx (only works if the sheet is "anyone with link")
    here = os.path.dirname(os.path.abspath(__file__))
    local_candidates = [
        os.path.join(here, "backend.xlsx"),
        os.path.join(here, "_backend_cache.xlsx"),
    ]
    for p in local_candidates:
        if os.path.exists(p):
            print(f"  Using local file: {p} ({os.path.getsize(p):,} bytes)")
            return p

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx"); tmp.close()

    # Try service-account auth if credentials are provided. The cron sets
    # GOOGLE_SERVICE_ACCOUNT_JSON from a GitHub secret; locally Ahmed can set
    # GOOGLE_SERVICE_ACCOUNT_PATH to a JSON key file.
    sa_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
    sa_path = os.environ.get("GOOGLE_SERVICE_ACCOUNT_PATH", "").strip()
    if sa_json or sa_path:
        try:
            import json as _json
            from google.oauth2 import service_account
            from google.auth.transport.requests import Request as _GAuthRequest
            info = _json.loads(sa_json) if sa_json else None
            scopes = ["https://www.googleapis.com/auth/drive.readonly"]
            if info:
                creds = service_account.Credentials.from_service_account_info(info, scopes=scopes)
            else:
                creds = service_account.Credentials.from_service_account_file(sa_path, scopes=scopes)
            creds.refresh(_GAuthRequest())
            # Drive v3 export endpoint converts a Google Sheet to xlsx server-side.
            url = (
                f"https://www.googleapis.com/drive/v3/files/{FILE_ID}/export"
                f"?mimeType=application%2Fvnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            print(f"  Fetching via service account (file: {FILE_ID})...")
            req = urllib.request.Request(url, headers={
                "Authorization": f"Bearer {creds.token}",
                "User-Agent": "hajj-ops/5.0",
            })
            with urllib.request.urlopen(req, timeout=60) as resp, open(tmp.name, "wb") as out:
                out.write(resp.read())
            print(f"  ✓ Downloaded {os.path.getsize(tmp.name):,} bytes (authenticated)")
            return tmp.name
        except Exception as e:
            print(f"  ⚠ Service-account auth failed: {e}")
            print(f"     Falling back to public export URL — likely 401 if sheet is private.")

    print(f"  Downloading from Google Drive (file: {FILE_ID})...")
    req = urllib.request.Request(DOWNLOAD_URL, headers={"User-Agent": "hajj-ops/5.0"})
    with urllib.request.urlopen(req, timeout=30) as resp, open(tmp.name, "wb") as out:
        out.write(resp.read())
    print(f"  ✓ Downloaded {os.path.getsize(tmp.name):,} bytes")
    return tmp.name

def read_sheet(wb, name, valid_first_col_pattern=None):
    if name not in wb.sheetnames:
        print(f"  ⚠ Sheet '{name}' not found")
        return []
    import re as _re
    ws = wb[name]
    headers = [s(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    rows = []
    pat = _re.compile(valid_first_col_pattern) if valid_first_col_pattern else None
    for r in range(2, ws.max_row + 1):
        v0 = ws.cell(r, 1).value
        if v0 is None: continue
        sv = s(v0)
        if sv == "" or sv.startswith("("): continue
        if sv.upper() == "TOTAL": continue
        if "Note:" in sv: continue
        if pat and not pat.match(sv): continue
        row = {}
        for c, h in enumerate(headers, 1):
            if h: row[h] = ws.cell(r, c).value
        rows.append(row)
    return rows

def parse_time_str(t):
    if isinstance(t, dtime): return t.hour
    if isinstance(t, str) and ":" in t: return int(t.split(":")[0])
    return None

def shift_covers_hour(start_t, end_t, target_hour):
    sh = parse_time_str(start_t); eh = parse_time_str(end_t)
    if sh is None or eh is None: return False
    if sh < eh: return sh <= target_hour < eh
    elif sh > eh: return target_hour >= sh or target_hour < eh
    else: return True

# ─── Compute personnel ────────────────────────────────────────────
def compute_personnel(roles_rows):
    by_role = {s(r.get("Role")): int(num(r.get("Target Count"))) for r in roles_rows if s(r.get("Role"))}
    leadership = sum(by_role.get(k, 0) for k in ["PM","Deputy PM","Admin Lead","Med Direction Lead"])
    paras = sum(v for k,v in by_role.items() if k not in ["PM","Deputy PM","Admin Lead","Med Direction Lead","GP"])
    gps = by_role.get("GP", 0)
    return {
        "total": leadership + paras + gps,
        "paramedics": paras, "gps": gps, "leadership_admin": leadership,
        "day_para": 127, "night_para": 124, "ambulances": 25,
        "sites": 10, "clinical_platforms": 9, "floating_functions": 3,
        "stations": 9, "clinics": 18,  # legacy keys kept for back-compat
    }

def _count_allocated_shifts(schedule_rows, units_rows):
    """Count total staff-shifts allocated across all DH days × shift slots."""
    unit_size = {s(u.get("Unit ID")): int(num(u.get("Size", 2))) for u in units_rows}
    total = 0
    for r in schedule_rows:
        uid = s(r.get("Unit ID"))
        if not uid: continue
        size = unit_size.get(uid, 2)
        for dh in range(4, 15):
            for slot in [1, 2]:
                code = s(r.get(f"{dh}DH-S{slot}"))
                if code: total += size
    return total

def compute_roster(staff_rows):
    LEADERSHIP_ROLES = ("PM","Deputy PM","Admin Lead","Med Direction Lead")
    para_rows = [r for r in staff_rows if s(r.get("Role")) not in LEADERSHIP_ROLES + ("GP",)]
    gp_rows = [r for r in staff_rows if s(r.get("Role")) == "GP"]
    leader_rows = [r for r in staff_rows if s(r.get("Role")) in LEADERSHIP_ROLES]
    def is_filled(r): return s(r.get("Status")) == "Filled" and s(r.get("Name")) != ""
    para_filled = sum(1 for r in para_rows if is_filled(r))
    gp_filled = sum(1 for r in gp_rows if is_filled(r))
    leader_filled = sum(1 for r in leader_rows if is_filled(r))
    para_total = len(para_rows); gp_total = len(gp_rows); leader_total = len(leader_rows)
    # all_total = paras + GPs + leadership/admin (full roster headline)
    all_total = para_total + gp_total + leader_total
    all_filled = para_filled + gp_filled + leader_filled
    by_subrole = Counter()
    by_subrole_status = defaultdict(lambda: {"vacant":0, "filled":0})
    ROLE_TO_SUB = {
        "Chief Paramedic":"Chief+Deputy","Deputy Chief Paramedic":"Chief+Deputy",
        "Logistics":"Logistics","Dispatch OCC":"Dispatch","SRCA Dispatch":"SRCA",
        "Depot":"Depot","Training":"Training","Supervisor":"Supervisor",
        "Ambulance Paramedic":"Alpha (Ambulance)","Foot Paramedic":"Romeo (Foot-Runner)","Medical Unit Paramedic":"Mike (Medical)",
    }
    for r in para_rows:
        role = s(r.get("Role"))
        sub = ROLE_TO_SUB.get(role, role)
        by_subrole[sub] += 1
        if is_filled(r): by_subrole_status[sub]["filled"] += 1
        else: by_subrole_status[sub]["vacant"] += 1
    return {
        "para_total":para_total,"para_vacant":para_total-para_filled,"para_filled":para_filled,
        "para_fill_pct":round(para_filled/max(para_total,1)*100,1),
        "gp_total":gp_total,"gp_vacant":gp_total-gp_filled,"gp_filled":gp_filled,
        "all_total":all_total,"all_vacant":all_total-all_filled,"all_filled":all_filled,
        "all_fill_pct":round(all_filled/max(all_total,1)*100,1),
        "by_subrole":dict(by_subrole),
        "by_subrole_status":{k:dict(v) for k,v in by_subrole_status.items()},
        "by_shift":{}, "by_home":{},
    }

def compute_org_structure(units_rows):
    out = []
    out.append({"category":"Leadership","role":"Project Manager","day":1,"night":0,"total":1,"notes":"Ahmed Alshudukhi MD"})
    out.append({"category":"Leadership","role":"Deputy PM","day":0,"night":1,"total":1,"notes":""})
    out.append({"category":"Leadership","role":"Admin Lead","day":1,"night":0,"total":1,"notes":""})
    out.append({"category":"Leadership","role":"Med Direction Lead","day":1,"night":0,"total":1,"notes":""})
    out.append({"category":"GP","role":"General Practitioners","day":10,"night":10,"total":20,"notes":"In Mike (Medical) units"})
    out.append({"category":"Para — CMD","role":"Chief Paramedic","day":1,"night":0,"total":1,"notes":""})
    out.append({"category":"Para — CMD","role":"Deputy Chief Paramedic","day":0,"night":1,"total":1,"notes":""})
    out.append({"category":"Para — Dispatch","role":"Dispatchers","day":2,"night":2,"total":4,"notes":""})
    out.append({"category":"Para — Depot","role":"Depot Clinic","day":2,"night":2,"total":4,"notes":""})
    out.append({"category":"Para — Logistics","role":"Logistics","day":4,"night":3,"total":7,"notes":"3 units sized [3,3,1]"})
    out.append({"category":"Para — Training","role":"Training","day":1,"night":1,"total":2,"notes":""})
    out.append({"category":"Para — Supervisor","role":"SUP-A (Arafat)","day":1,"night":1,"total":2,"notes":""})
    out.append({"category":"Para — Supervisor","role":"SUP-Z (Muzdalifah)","day":1,"night":1,"total":2,"notes":""})
    out.append({"category":"Para — Supervisor","role":"SUP-M (Mina)","day":1,"night":1,"total":2,"notes":""})
    by_station = defaultdict(lambda: {"M":0,"A":0,"R":0})
    for u in units_rows:
        utype = s(u.get("Unit Type"))
        home = s(u.get("Home Station"))
        if utype == "Medical": by_station[home]["M"] += 1
        elif utype == "Ambulance": by_station[home]["A"] += 1
        elif utype == "Foot-Runner": by_station[home]["R"] += 1
    for st in STATIONS:
        m = by_station[st]
        total_units = m["M"] + m["A"] + m["R"]
        total_paras = m["M"]*2 + m["A"]*2 + m["R"]*2
        out.append({"category":"Para — Station","role":st,"day":total_units,"night":total_units,
                    "total":total_paras,"notes":f"{m['M']} Mike + {m['A']} Alpha + {m['R']} Romeo = {total_units} units"})
    return out

def build_hourly(staff_rows, schedule_rows, shifts_rows, units_rows, ambulance_rows=None):
    shift_map = {s(r.get("Code")): (r.get("Start"), r.get("End")) for r in shifts_rows if s(r.get("Code"))}
    unit_size = {s(u.get("Unit ID")): int(num(u.get("Size", 2))) for u in units_rows}
    unit_home = {s(u.get("Unit ID")): s(u.get("Home Station")) for u in units_rows}
    unit_type = {s(u.get("Unit ID")): s(u.get("Unit Type")) for u in units_rows}
    schedule = defaultdict(lambda: defaultdict(list))
    for r in schedule_rows:
        uid = s(r.get("Unit ID"))
        if not uid: continue
        for dh in range(4, 15):
            for slot in [1, 2]:
                code = s(r.get(f"{dh}DH-S{slot}"))
                if code: schedule[uid][dh].append(code)

    # Pre-compute roster totals (sum of unit sizes) overall and by zone.
    # Used by the positioning page to display "Off Duty" counts.
    def _zone_of(home):
        if home.startswith("ARF"): return "Arafat"
        if home.startswith("MUZ"): return "Muzdalifah"
        if home.startswith("MIN"): return "Mina"
        return "Support"
    total_roster_by_zone = {"Arafat":0, "Muzdalifah":0, "Mina":0, "Support":0}
    for uid, sz in unit_size.items():
        total_roster_by_zone[_zone_of(unit_home.get(uid, ""))] += sz
    total_roster = sum(total_roster_by_zone.values())

    # Index ambulances by home station for hourly active-amb counts
    amb_by_home = defaultdict(list)
    for a in (ambulance_rows or []):
        home = s(a.get("Home Station"))
        if not home: continue
        amb_by_home[home].append({
            "id": s(a.get("Ambulance ID")),
            "day_crew": s(a.get("Day Alpha Crew")),
            "night_crew": s(a.get("Night Alpha Crew")),
        })

    # Sites that get per-station detail (9 clinical + OCC depot).
    SITES = STATIONS + ["OCC"]

    # Static physical ambulance count per site: ambulances are parked at
    # their home station 24/7, independent of which crew is currently on
    # duty. The earlier "crewed count" hid 5+ ambulances during shift gaps,
    # which was misleading — the vehicles never left.
    static_amb_by_home = {st: 0 for st in SITES}
    for st, ambs in amb_by_home.items():
        if st in static_amb_by_home:
            static_amb_by_home[st] = len(ambs)

    DH_RANGE = list(range(4, 15))  # DH 4 through DH 14 inclusive
    out_hours = []
    for dh in DH_RANGE:
        # DH 4-6: medical pre-mob, day-hours only. DH 7-14: full 24h
        # (Movement A regular metro starts DH 7 08:00 and runs to DH 8 00:00).
        start_hour = 6 if dh in [4,5,6] else 0
        end_hour = 18 if dh in [4,5,6] else 24
        for h in range(start_hour, end_hour):
            # Movement label — strict membership: an hour is labeled with a
            # movement ONLY when it falls inside that movement's declared
            # window [start, end). Hours outside any defined window get "—"
            # (not GAP — GAP is itself an explicit movement window). This
            # matches operational reality: gaps between movements (e.g.
            # 11:00-12:00 on DH 9, 20:00-21:00 on DH 9) are TRANSITION
            # periods, not part of the prior or next phase.
            cur = dh*100 + h
            mvt_code = "—"; shift_label = "DAY" if 6 <= h < 18 else "NIGHT"
            for ph in MOVEMENT_PHASES:
                sd = int(ph["start_dh"].split()[0]); ed = int(ph["end_dh"].split()[0])
                sh_p = int(ph["start_hour"].split(":")[0]); eh_p = int(ph["end_hour"].split(":")[0])
                start = sd*100 + sh_p; end = ed*100 + eh_p
                if start <= cur < end:
                    mvt_code = ph["mvt"]; shift_label = ph["shift"]
                    break

            zones = {"Arafat":0, "Muzdalifah":0, "Mina":0, "Support":0}
            stations = {st:0 for st in SITES}
            by_type = defaultdict(int)
            by_zone_type = {
                "Arafat": defaultdict(int),
                "Muzdalifah": defaultdict(int),
                "Mina": defaultdict(int),
                "Support": defaultdict(int),
            }
            active_units = set()
            # Per-site detail: track everything operationally useful.
            sd_paras       = {st: 0 for st in SITES}
            sd_by_type     = {st: defaultdict(int) for st in SITES}
            sd_units       = {st: [] for st in SITES}
            sd_shifts      = {st: set() for st in SITES}
            sd_doctors     = {st: 0 for st in SITES}

            for uid, day_shifts in schedule.items():
                if dh not in day_shifts: continue
                matched_shift = None
                for code in day_shifts[dh]:
                    if code in shift_map:
                        start_t, end_t = shift_map[code]
                        if shift_covers_hour(start_t, end_t, h):
                            matched_shift = code
                            break
                if not matched_shift: continue
                active_units.add(uid)
                size = unit_size.get(uid, 2)
                home = unit_home.get(uid, "")
                utype = unit_type.get(uid, "") or "Other"
                by_type[utype] += 1
                zone_name = _zone_of(home)
                by_zone_type[zone_name][utype] += 1
                if home.startswith("ARF"):
                    zones["Arafat"] += size
                elif home.startswith("MUZ"):
                    zones["Muzdalifah"] += size
                elif home.startswith("MIN"):
                    zones["Mina"] += size
                else:
                    zones["Support"] += size
                if home in stations:
                    stations[home] += size
                    sd_paras[home] += size
                    sd_by_type[home][utype] += 1
                    sd_units[home].append(uid)
                    sd_shifts[home].add(matched_shift)
                    if utype == "Doctor":
                        sd_doctors[home] += 1

            # Crewed ambulance presence at each station — the "currently
            # has a crew on duty" view. This is what `stations_amb` shows
            # by default. The static physical count is kept as a secondary
            # `stations_amb_total` field so the detail panel can show "x/y".
            stations_amb = {st: 0 for st in SITES}
            for st, ambs in amb_by_home.items():
                if st not in stations_amb: continue
                for a in ambs:
                    day_crew = a["day_crew"]; night_crew = a["night_crew"]
                    if (day_crew and day_crew in active_units) or (night_crew and night_crew in active_units):
                        stations_amb[st] += 1

            stations_amb_total = dict(static_amb_by_home)

            # Doctors per station: derived from active Doctor-type units
            # (Delta-*). Each station has 2 Delta units (day GP slot + night
            # GP slot) — sd_doctors picks up both when both are on duty,
            # which is the correct picture during handover overlap hours.
            stations_doctors = dict(sd_doctors)
            total_doctors = sum(stations_doctors.values())

            # Build per-site detail blob for the positioning view's expander.
            stations_detail = {}
            for st in SITES:
                stations_detail[st] = {
                    "paras":          sd_paras[st],
                    "by_type":        dict(sd_by_type[st]),
                    "doctors":        stations_doctors[st],
                    "amb_crewed":     stations_amb[st],
                    "amb_total":      stations_amb_total.get(st, 0),
                    "active_units":   sorted(sd_units[st]),
                    "active_shifts":  sorted(sd_shifts[st]),
                }

            arf_a = stations_amb.get("ARF1",0)+stations_amb.get("ARF2",0)+stations_amb.get("ARF3",0)
            muz_a = stations_amb.get("MUZ1",0)+stations_amb.get("MUZ2",0)+stations_amb.get("MUZ3",0)
            min_a = stations_amb.get("MIN1",0)+stations_amb.get("MIN2",0)+stations_amb.get("MIN3",0)
            grand_a = sum(stations_amb.values())
            grand_a_total = sum(stations_amb_total.values())

            out_hours.append({
                "dh":f"{dh} DH","hour":f"{h:02d}:00","mvt":mvt_code,"shift":shift_label,
                "label":f"{dh} DH {h:02d}:00",
                "arf_s":zones["Arafat"],"muz_s":zones["Muzdalifah"],"min_s":zones["Mina"],
                "arf_a":arf_a,"muz_a":muz_a,"min_a":min_a,
                "rov_c":0,"fwd_c":0,"dep_c":0,
                "support":zones["Support"],"rov_a":0,"fwd_a":0,"dep_a":0,
                "stations":stations, "stations_amb":stations_amb,
                "stations_amb_total": stations_amb_total,
                "stations_doctors":  stations_doctors,
                "stations_detail":   stations_detail,
                "doctors":           total_doctors,
                "grand_s":sum(zones.values()),"grand_a":grand_a,
                "grand_a_total":     grand_a_total,
                "by_type": dict(by_type),
                "by_zone_type": {z: dict(d) for z, d in by_zone_type.items()},
                "units_active": len(active_units),
            })

    return {
        "hours":out_hours,
        "peak_arafat":max((h["arf_s"] for h in out_hours), default=0),
        "peak_muzdalifah":max((h["muz_s"] for h in out_hours), default=0),
        "peak_mina":max((h["min_s"] for h in out_hours), default=0),
        "movement_peaks":{},"total_hours":len(out_hours),
        "total_roster": total_roster,
        "total_roster_by_zone": total_roster_by_zone,
    }

def compute_augmentations(aug_rows):
    if not aug_rows:
        return {"total":0,"active":0,"planned":0,"returned":0,"cancelled":0,"total_para_moved":0,
                "dominant_status":"None","by_movement":{},"by_donor":{},"by_recipient":{},
                "matrix":{},"sample":[]}
    by_status = Counter(s(r.get("Status")) for r in aug_rows)
    # Build donor → recipient matrix, total + per-movement
    matrix = {"all": defaultdict(lambda: defaultdict(int))}
    for r in aug_rows:
        donor_unit = s(r.get("From Unit"))
        # Extract station from unit ID (e.g. "Alpha-25" → look up home, but for sheet we have raw station codes too)
        donor_st = s(r.get("From Station") or r.get("From"))
        if not donor_st:
            # Best-effort: parse from unit prefix or fallback to unit name
            donor_st = donor_unit
        recipient = s(r.get("To Station") or r.get("To"))
        mvt = s(r.get("Movement"))
        paras = int(num(r.get("Paras", 1)))
        if not donor_st or not recipient: continue
        matrix["all"][donor_st][recipient] += paras
        matrix.setdefault(mvt, defaultdict(lambda: defaultdict(int)))
        matrix[mvt][donor_st][recipient] += paras
    # Convert defaultdicts to plain dicts for JSON
    matrix_json = {k: {d: dict(rs) for d, rs in v.items()} for k, v in matrix.items()}
    return {
        "total":len(aug_rows),
        "active":by_status.get("Active",0),"planned":by_status.get("Planned",0),
        "returned":by_status.get("Returned",0),"cancelled":by_status.get("Cancelled",0),
        "total_para_moved":sum(int(num(r.get("Paras",1))) for r in aug_rows),
        "dominant_status":by_status.most_common(1)[0][0] if by_status else "None",
        "by_movement":dict(Counter(s(r.get("Movement")) for r in aug_rows)),
        "by_donor":dict(Counter(s(r.get("From Unit")) for r in aug_rows)),
        "by_recipient":dict(Counter(s(r.get("To Station")) for r in aug_rows)),
        "matrix": matrix_json,
        "sample":aug_rows[:24],
    }

def compute_status_counts(units_rows):
    return {"SUPPORT":98,"SURGE":52,"ACTIVE":38,"STANDBY":36}

def compute_schedule_grid(schedule_rows, units_rows, shifts_rows):
    """Per-station per-day-slot summary: how many units & paras on each shift."""
    shift_dur = {s(r.get("Code")): compute_shift_duration(r.get("Start"), r.get("End")) for r in shifts_rows if s(r.get("Code"))}
    unit_size = {s(u.get("Unit ID")): int(num(u.get("Size", 1))) for u in units_rows}
    unit_home = {s(u.get("Unit ID")): s(u.get("Home Station")) for u in units_rows}
    unit_type = {s(u.get("Unit ID")): s(u.get("Unit Type")) for u in units_rows}
    DH_DAYS = list(range(4, 15))
    grid = {}
    for dh in DH_DAYS:
        grid[dh] = {"slot1": defaultdict(lambda: {"units":0,"paras":0,"by_type":Counter()}),
                    "slot2": defaultdict(lambda: {"units":0,"paras":0,"by_type":Counter()})}
    for r in schedule_rows:
        uid = s(r.get("Unit ID"))
        if not uid: continue
        size = unit_size.get(uid, 1)
        home = unit_home.get(uid, "")
        utype = unit_type.get(uid, "")
        for dh in DH_DAYS:
            for slot_num, slot_key in [(1,"slot1"),(2,"slot2")]:
                code = s(r.get(f"{dh}DH-S{slot_num}"))
                if code:
                    cell = grid[dh][slot_key][home]
                    cell["units"] += 1
                    cell["paras"] += size
                    cell["by_type"][utype] += 1
    out = {}
    for dh in DH_DAYS:
        out[str(dh)] = {
            "slot1": {st: {"units": d["units"], "paras": d["paras"], "by_type": dict(d["by_type"])} for st, d in grid[dh]["slot1"].items()},
            "slot2": {st: {"units": d["units"], "paras": d["paras"], "by_type": dict(d["by_type"])} for st, d in grid[dh]["slot2"].items()},
        }
    return out

def compute_daily_view(schedule_rows, units_rows, shifts_rows):
    """One row per (DH day, slot) with totals. For day-by-day overview."""
    shift_dur = {s(r.get("Code")): compute_shift_duration(r.get("Start"), r.get("End")) for r in shifts_rows if s(r.get("Code"))}
    shift_type = {s(r.get("Code")): s(r.get("Type")) for r in shifts_rows if s(r.get("Code"))}
    unit_size = {s(u.get("Unit ID")): int(num(u.get("Size", 1))) for u in units_rows}
    unit_home = {s(u.get("Unit ID")): s(u.get("Home Station")) for u in units_rows}
    unit_type = {s(u.get("Unit ID")): s(u.get("Unit Type")) for u in units_rows}
    DH_DAYS = list(range(4, 15))
    out = []
    for dh in DH_DAYS:
        day_units = set()
        day_paras = 0
        zones = {"Arafat":0,"Muzdalifah":0,"Mina":0,"Support":0}
        types = {"Medical":0,"Ambulance":0,"Foot-Runner":0,"Other":0}
        shift_breakdown = Counter()
        day_total_hours = 0
        for r in schedule_rows:
            uid = s(r.get("Unit ID"))
            if not uid: continue
            size = unit_size.get(uid, 1)
            home = unit_home.get(uid, "")
            utype = unit_type.get(uid, "")
            unit_active_today = False
            for slot in [1,2]:
                code = s(r.get(f"{dh}DH-S{slot}"))
                if code:
                    unit_active_today = True
                    shift_breakdown[code] += size
                    day_total_hours += size * shift_dur.get(code, 0)
            if unit_active_today:
                day_units.add(uid)
                day_paras += size
                if home.startswith("ARF"): zones["Arafat"] += size
                elif home.startswith("MUZ"): zones["Muzdalifah"] += size
                elif home.startswith("MIN"): zones["Mina"] += size
                else: zones["Support"] += size
                if utype in types: types[utype] += size
                else: types["Other"] += size
        out.append({
            "dh": dh,
            "label": f"{dh} DH",
            "active_units": len(day_units),
            "active_paras": day_paras,
            "zones": zones,
            "types": types,
            "shifts": dict(shift_breakdown.most_common()),
            "total_hours": int(day_total_hours),
        })
    return out

def compute_status_live(staff_rows, schedule_rows):
    """Real status mix from filled/vacant/scheduled data."""
    by_status = Counter()
    by_role_status = defaultdict(lambda: Counter())
    for r in staff_rows:
        status = s(r.get("Status")) or "Vacant"
        role = s(r.get("Role"))
        if not role: continue
        by_status[status] += 1
        by_role_status[role][status] += 1
    # Also schedule-based: which units have any shift assigned
    units_with_schedule = set()
    for r in schedule_rows:
        uid = s(r.get("Unit ID"))
        if not uid: continue
        for dh in range(4,15):
            for slot in [1,2]:
                if s(r.get(f"{dh}DH-S{slot}")):
                    units_with_schedule.add(uid)
                    break
    return {
        "by_status": dict(by_status),
        "by_role_status": {k: dict(v) for k, v in by_role_status.items()},
        "units_scheduled": len(units_with_schedule),
    }

def compute_zone_movement(schedule_rows, units_rows, shifts_rows):
    """Zone × Movement matrix using MOVEMENT_PHASES timing.
       Handles night shifts that wrap past midnight — a NIGHT shift assigned to
       column NDH-S2 covers hours 18..23 on day N AND hours 00..05 on day N+1.
       So when checking hour H on day D, we test:
         - DDH-S1 / DDH-S2 (assignments scheduled on day D)
         - if H < 12, also (D-1)DH-S2 (a night shift from previous day extends here)"""
    unit_size = {s(u.get("Unit ID")): int(num(u.get("Size", 1))) for u in units_rows}
    unit_home = {s(u.get("Unit ID")): s(u.get("Home Station")) for u in units_rows}
    shift_map = {s(r.get("Code")): (r.get("Start"), r.get("End")) for r in shifts_rows if s(r.get("Code"))}
    out = {}
    for ph in MOVEMENT_PHASES:
        zones = {"Arafat":0,"Muzdalifah":0,"Mina":0,"Support":0}
        sd = int(ph["start_dh"].split()[0]); ed = int(ph["end_dh"].split()[0])
        sh_p = int(ph["start_hour"].split(":")[0]); eh_p = int(ph["end_hour"].split(":")[0])
        # Build absolute (dh,hour) list covering this phase
        phase_slots = []
        cur_dh, cur_h = sd, sh_p
        end_dh_h = ed * 24 + eh_p
        while cur_dh * 24 + cur_h <= end_dh_h:
            phase_slots.append((cur_dh, cur_h))
            cur_h += 1
            if cur_h >= 24:
                cur_h = 0; cur_dh += 1
            if len(phase_slots) > 500: break  # safety
        for r in schedule_rows:
            uid = s(r.get("Unit ID"))
            if not uid: continue
            home = unit_home.get(uid, "")
            size = unit_size.get(uid, 1)
            unit_active = False
            for dh_p, h_p in phase_slots:
                if unit_active: break
                # Build candidate (col_dh, slot) pairs to test
                candidates = [(dh_p, 1), (dh_p, 2)]
                if h_p < 12:  # early-morning hour might belong to prior day's night shift
                    candidates.append((dh_p - 1, 2))
                for col_dh, slot in candidates:
                    code = s(r.get(f"{col_dh}DH-S{slot}"))
                    if not code or code not in shift_map: continue
                    start_t, end_t = shift_map[code]
                    if shift_covers_hour(start_t, end_t, h_p):
                        unit_active = True
                        break
            if unit_active:
                if home.startswith("ARF"): zones["Arafat"] += size
                elif home.startswith("MUZ"): zones["Muzdalifah"] += size
                elif home.startswith("MIN"): zones["Mina"] += size
                else: zones["Support"] += size
        out[ph["mvt"]] = zones
    return out

def compute_zone_day(schedule_rows, units_rows):
    """Zone × Day matrix (paras per zone per DH day)."""
    unit_size = {s(u.get("Unit ID")): int(num(u.get("Size", 1))) for u in units_rows}
    unit_home = {s(u.get("Unit ID")): s(u.get("Home Station")) for u in units_rows}
    out = {}
    for dh in range(4, 15):
        zones = {"Arafat":0,"Muzdalifah":0,"Mina":0,"Support":0}
        for r in schedule_rows:
            uid = s(r.get("Unit ID"))
            if not uid: continue
            active = False
            for slot in [1,2]:
                if s(r.get(f"{dh}DH-S{slot}")):
                    active = True; break
            if not active: continue
            size = unit_size.get(uid, 1)
            home = unit_home.get(uid, "")
            if home.startswith("ARF"): zones["Arafat"] += size
            elif home.startswith("MUZ"): zones["Muzdalifah"] += size
            elif home.startswith("MIN"): zones["Mina"] += size
            else: zones["Support"] += size
        out[str(dh)] = zones
    return out

def compute_role_views(units_detail, staff_rows, ambulance_roster):
    """Pre-computed summaries for each role view."""
    pm_view = {
        "total_units": len(units_detail),
        "total_paras": sum(int(num(r.get("Slot")) if isinstance(r.get("Slot"),(int,float)) else 1) for r in staff_rows if s(r.get("Role")) not in ("PM","Deputy PM","Admin Lead","Med Direction Lead","GP")),
        "filled_pct": 0,
    }
    chief_view = {
        "command_units": [u for u in units_detail if u["category"] == "Command"],
        "leadership": [u for u in units_detail if u["category"] == "Leadership"],
    }
    supervisor_view = {
        "stations": {st: [u for u in units_detail if u["home"] == st] for st in
                     ["ARF1","ARF2","ARF3","MUZ1","MUZ2","MUZ3","MIN1","MIN2","MIN3"]},
    }
    paramedic_view = {
        "operational_units": [u for u in units_detail if u["category"] == "Operational"],
        "ambulances": ambulance_roster,
    }
    executive_view = {
        "headlines": {
            "people": sum(u["size"] for u in units_detail),
            "units": len(units_detail),
            "sites": 10,
            "clinical_platforms": 9,
            "stations": 9,  # legacy back-compat
            "ambulances": len(ambulance_roster),
        },
    }
    return {
        "pm": pm_view,
        "chief": chief_view,
        "supervisor": supervisor_view,
        "paramedic": paramedic_view,
        "executive": executive_view,
    }

def compute_units_detail(units_rows, staff_rows):
    members_by_unit = defaultdict(list)
    for r in staff_rows:
        unit = s(r.get("Unit"))
        if not unit: continue
        members_by_unit[unit].append({
            "staff_id": s(r.get("Staff ID")),
            "name": s(r.get("Name")),
            "role": s(r.get("Role")),
            "slot": s(r.get("Slot")),
            "phone": s(r.get("Phone")),
            "email": s(r.get("Email")),
            "call_sign": s(r.get("Radio Call Sign")) or unit,
            "status": s(r.get("Status")),
        })
    out = []
    for u in units_rows:
        uid = s(u.get("Unit ID"))
        if not uid: continue
        members = members_by_unit.get(uid, [])
        out.append({
            "id": uid,
            "type": s(u.get("Unit Type")),
            "category": s(u.get("Category")),
            "size": int(num(u.get("Size", 1))),
            "home": s(u.get("Home Station")),
            "default_shift": s(u.get("Default Shift")),
            "tags": s(u.get("Tags")),
            "notes": s(u.get("Notes")),
            "members": members,
            "filled_count": sum(1 for m in members if m["status"] == "Filled" and m["name"]),
            "total_count": len(members),
        })
    return out


def compute_accommodation_live(units_rows, ambulance_rows):
    """Derive accommodation from Home Station column. Each person needs a bed."""
    unit_size = {s(u.get("Unit ID")): int(num(u.get("Size", 1))) for u in units_rows}
    unit_home = {s(u.get("Unit ID")): s(u.get("Home Station")) for u in units_rows}
    by_home = defaultdict(lambda: {"sta_para":0,"amb_crew":0,"gps":0,"support":0})
    # From Units
    for u in units_rows:
        uid = s(u.get("Unit ID"))
        if not uid: continue
        home = unit_home.get(uid, "")
        size = unit_size.get(uid, 1)
        utype = s(u.get("Unit Type"))
        cat = s(u.get("Category"))
        if not home: continue
        if utype == "Ambulance":
            by_home[home]["amb_crew"] += size
        elif utype == "Medical":
            # Mike units include 1 GP, 2 paras
            by_home[home]["sta_para"] += max(0, size - 1)
            by_home[home]["gps"] += 1
        elif cat == "Operational":
            by_home[home]["sta_para"] += size
        else:
            by_home[home]["support"] += size
    # Build output rows
    out = []
    for home, d in sorted(by_home.items()):
        total = d["sta_para"] + d["amb_crew"] + d["gps"] + d["support"]
        out.append({
            "location": home,
            "sta_para": d["sta_para"],
            "amb_crew": d["amb_crew"],
            "gps": d["gps"],
            "support": d["support"],
            "rov_fwd": 0,
            "total_beds": total,
            "bunk_sets": (total + 1) // 2,
        })
    return out

def compute_ambulance_dashboards(amb_rows, units_rows, schedule_rows, shifts_rows):
    """Four ambulance analytics:
       a) per-ambulance scheduled hours across 11 days
       b) availability per day
       c) day vs night ambulance distribution by zone
       d) crew assignment Gantt rows
    """
    unit_home = {s(u.get("Unit ID")): s(u.get("Home Station")) for u in units_rows}
    schedule_by_uid = {s(r.get("Unit ID")): r for r in schedule_rows if s(r.get("Unit ID"))}
    shift_dur = {s(r.get("Code")): compute_shift_duration(r.get("Start"), r.get("End")) for r in shifts_rows if s(r.get("Code"))}
    shift_type = {s(r.get("Code")): s(r.get("Type")) for r in shifts_rows if s(r.get("Code"))}

    DH_DAYS = list(range(4, 15))

    # a) Per-ambulance scheduled hours: ambulance is scheduled when its day OR night Alpha unit is on shift
    amb_hours = {}  # {amb_id: {dh: hours}}
    amb_meta = {}
    for r in amb_rows:
        aid = s(r.get("Ambulance ID"))
        if not aid: continue
        day_alpha = s(r.get("Day Alpha Crew"))
        night_alpha = s(r.get("Night Alpha Crew"))
        atype = s(r.get("Type"))
        home = s(r.get("Home Station")) or unit_home.get(day_alpha, "")
        amb_meta[aid] = {"type": atype, "home": home}
        amb_hours[aid] = {dh: 0 for dh in DH_DAYS}
        for crew_unit in [day_alpha, night_alpha]:
            if not crew_unit or crew_unit not in schedule_by_uid: continue
            row = schedule_by_uid[crew_unit]
            for dh in DH_DAYS:
                for slot in [1, 2]:
                    code = s(row.get(f"{dh}DH-S{slot}"))
                    if code:
                        amb_hours[aid][dh] += shift_dur.get(code, 0)

    # b) Availability per day: count of ambulances active on any given day
    daily_active = {dh: 0 for dh in DH_DAYS}
    daily_by_type = {dh: {"Essential":0,"Backup":0,"Roving":0} for dh in DH_DAYS}
    for aid, by_dh in amb_hours.items():
        atype = amb_meta[aid]["type"]
        for dh, hrs in by_dh.items():
            if hrs > 0:
                daily_active[dh] += 1
                if atype in daily_by_type[dh]:
                    daily_by_type[dh][atype] += 1

    # c) Day vs Night distribution by zone
    # For each ambulance, determine if its day_alpha is scheduled on a DAY shift, or night_alpha on NIGHT shift
    day_by_zone = {"Arafat":0,"Muzdalifah":0,"Mina":0,"Support":0}
    night_by_zone = {"Arafat":0,"Muzdalifah":0,"Mina":0,"Support":0}
    for r in amb_rows:
        aid = s(r.get("Ambulance ID"))
        if not aid: continue
        day_alpha = s(r.get("Day Alpha Crew"))
        night_alpha = s(r.get("Night Alpha Crew"))
        home = s(r.get("Home Station")) or unit_home.get(day_alpha, "")
        zone = ("Arafat" if home.startswith("ARF") else
                "Muzdalifah" if home.startswith("MUZ") else
                "Mina" if home.startswith("MIN") else "Support")
        # Day crew exists?
        if day_alpha and day_alpha in schedule_by_uid:
            day_by_zone[zone] += 1
        if night_alpha and night_alpha in schedule_by_uid:
            night_by_zone[zone] += 1

    # d) Gantt-style: list per-ambulance shift segments
    amb_to_crews = {}  # aid -> (day_crew, night_crew)
    for r in amb_rows:
        aid = s(r.get("Ambulance ID"))
        if aid:
            amb_to_crews[aid] = (s(r.get("Day Alpha Crew")), s(r.get("Night Alpha Crew")))
    gantt = []
    for aid in sorted(amb_hours.keys()):
        meta = amb_meta[aid]
        day_crew, night_crew = amb_to_crews.get(aid, ("", ""))
        segments = []
        for crew_role, crew_unit in [("Day", day_crew), ("Night", night_crew)]:
            if not crew_unit or crew_unit not in schedule_by_uid: continue
            row = schedule_by_uid[crew_unit]
            for dh in DH_DAYS:
                for slot in [1,2]:
                    code = s(row.get(f"{dh}DH-S{slot}"))
                    if code:
                        segments.append({"dh": dh, "slot": slot, "code": code, "role": crew_role,
                                         "duration": shift_dur.get(code, 0), "type": shift_type.get(code, "")})
        gantt.append({"id": aid, "type": meta["type"], "home": meta["home"],
                      "day_crew": day_crew, "night_crew": night_crew, "segments": segments})

    return {
        "amb_hours": amb_hours,
        "amb_meta": amb_meta,
        "daily_active": daily_active,
        "daily_by_type": daily_by_type,
        "day_by_zone": day_by_zone,
        "night_by_zone": night_by_zone,
        "gantt": gantt[:60],  # cap for size
    }

def compute_day_night_per_station(units_rows, schedule_rows, shifts_rows):
    """For each station: day-shift paras/units count vs night-shift paras/units count."""
    shift_type = {s(r.get("Code")): s(r.get("Type")) for r in shifts_rows if s(r.get("Code"))}
    unit_size = {s(u.get("Unit ID")): int(num(u.get("Size", 1))) for u in units_rows}
    unit_home = {s(u.get("Unit ID")): s(u.get("Home Station")) for u in units_rows}
    out = defaultdict(lambda: {"day_paras": 0, "night_paras": 0, "day_units": set(), "night_units": set()})
    for r in schedule_rows:
        uid = s(r.get("Unit ID"))
        if not uid: continue
        home = unit_home.get(uid, "")
        size = unit_size.get(uid, 1)
        if not home: continue
        for dh in range(4, 15):
            for slot in [1,2]:
                code = s(r.get(f"{dh}DH-S{slot}"))
                if not code: continue
                stype = shift_type.get(code, "").upper()
                if "NIGHT" in stype:
                    out[home]["night_paras"] += size
                    out[home]["night_units"].add(uid)
                else:
                    out[home]["day_paras"] += size
                    out[home]["day_units"].add(uid)
    return {st: {"day_paras": d["day_paras"], "night_paras": d["night_paras"],
                 "day_units": len(d["day_units"]), "night_units": len(d["night_units"])}
            for st, d in out.items()}


def compute_ambulance_data(amb_rows, units_rows):
    """Build ambulance roster table + per-station counts (total + by type).
       Home Station is computed from Day Alpha crew's home (mirrors xlsx formula)."""
    unit_home = {s(u.get("Unit ID")): s(u.get("Home Station")) for u in units_rows}
    roster = []
    by_station = defaultdict(int)
    by_station_type = defaultdict(lambda: {"Essential": 0, "Backup": 0, "Roving": 0})
    for r in amb_rows:
        aid = s(r.get("Ambulance ID"))
        if not aid: continue
        day = s(r.get("Day Alpha Crew"))
        night = s(r.get("Night Alpha Crew"))
        atype = s(r.get("Type"))
        home = s(r.get("Home Station"))
        if not home and day:
            home = unit_home.get(day, "")
        roster.append({
            "id": aid, "type": atype, "cls": s(r.get("Class")),
            "day_crew": day, "night_crew": night, "home": home,
            "status": s(r.get("Status")) or "Ready", "notes": s(r.get("Notes")),
        })
        if home:
            by_station[home] += 1
            if atype in by_station_type[home]:
                by_station_type[home][atype] += 1
    return roster, dict(by_station), {k: dict(v) for k, v in by_station_type.items()}

def compute_stations_detail(units_rows, staff_rows, unit_readiness_rows):
    unit_filled = defaultdict(lambda: {"total":0,"filled":0})
    for r in staff_rows:
        unit = s(r.get("Unit"))
        if not unit: continue
        unit_filled[unit]["total"] += 1
        if s(r.get("Status")) == "Filled" and s(r.get("Name")):
            unit_filled[unit]["filled"] += 1
    unit_ready = {}
    for r in unit_readiness_rows:
        uid = s(r.get("Unit ID"))
        if uid: unit_ready[uid] = s(r.get("Ready"))
    by_home = defaultdict(list)
    for u in units_rows:
        uid = s(u.get("Unit ID"))
        utype = s(u.get("Unit Type"))
        cat = s(u.get("Category"))
        size = int(num(u.get("Size", 2)))
        home = s(u.get("Home Station"))
        if not home: continue
        f = unit_filled[uid]
        by_home[home].append({
            "id": uid, "type": utype, "category": cat, "size": size,
            "filled": f["filled"], "total": f["total"],
            "ready": unit_ready.get(uid, "")
        })
    out = {}
    for st, units in by_home.items():
        by_type = Counter(u["type"] for u in units)
        out[st] = {
            "units": sorted(units, key=lambda u: (u["category"] != "Operational", u["type"], u["id"])),
            "by_type": dict(by_type),
            "total_units": len(units),
            "total_size": sum(u["size"] for u in units),
            "total_filled_paras": sum(u["filled"] for u in units),
        }
    return out

# ─── Main ─────────────────────────────────────────────────────────
def compute_org_tree(units_rows, staff_rows):
    """Build hierarchical reporting tree.
       PM → DPM/ADM/MDL → CHF → DCH → SUP-{A,Z,M} → ops units at each station."""
    # Index units by id
    by_id = {s(u.get("Unit ID")): u for u in units_rows}
    home_groups = defaultdict(list)
    for u in units_rows:
        uid = s(u.get("Unit ID"))
        home_groups[s(u.get("Home Station"))].append(uid)
    # Index staff by unit
    members = defaultdict(list)
    for r in staff_rows:
        unit = s(r.get("Unit"))
        if unit:
            members[unit].append({
                "name": s(r.get("Name")), "role": s(r.get("Role")),
                "status": s(r.get("Status")), "call_sign": s(r.get("Radio Call Sign")) or unit
            })
    def node(uid, label=None):
        u = by_id.get(uid, {})
        m = members.get(uid, [])
        filled = sum(1 for x in m if x.get("status") == "Filled" and x.get("name"))
        size = int(num(u.get("Size", 1)))
        return {
            "id": uid, "label": label or uid,
            "type": s(u.get("Unit Type")), "category": s(u.get("Category")),
            "home": s(u.get("Home Station")),
            "size": size, "filled": filled,
            "members": m,
            "children": [],
        }
    # Build tree
    root = node("PM", "Project Manager")
    # Tier 1: leadership peer to PM
    for uid in ["DPM", "ADM", "MDL"]:
        if uid in by_id:
            root["children"].append(node(uid))
    # Tier 2: CHF under PM
    chf = node("CHF") if "CHF" in by_id else None
    if chf:
        root["children"].append(chf)
        # DCH under CHF
        if "DCH" in by_id:
            dch = node("DCH")
            chf["children"].append(dch)
            # Supervisors organized by zone
            zone_groups = {
                "Arafat": [u for u in ["SUP-A1", "SUP-A2"] if u in by_id],
                "Muzdalifah": [u for u in ["SUP-Z1", "SUP-Z2"] if u in by_id],
                "Mina": [u for u in ["SUP-M1", "SUP-M2"] if u in by_id],
            }
            for zone, sup_ids in zone_groups.items():
                for sup_id in sup_ids:
                    sup = node(sup_id)
                    dch["children"].append(sup)
                    # Units at this supervisor's home station
                    sup_home = by_id.get(sup_id, {}).get("Home Station", "")
                    sup_home = s(sup_home)
                    for uid in home_groups.get(sup_home, []):
                        u = by_id.get(uid, {})
                        if s(u.get("Category")) == "Operational":
                            sup["children"].append(node(uid))
        # Support functions under CHF (separate branch)
        for uid in ["LOG-1", "LOG-2", "LOG-3", "OCC-1", "OCC-2", "SRCA-1", "TRN-1"]:
            if uid in by_id:
                chf["children"].append(node(uid))
    # Catch-all: any remaining units not yet placed (e.g. ARF3 ops not under any SUP)
    placed = set()
    def collect(n):
        placed.add(n["id"])
        for c in n["children"]: collect(c)
    collect(root)
    unplaced_branch = node("UNASSIGNED", "Other Units")
    for u in units_rows:
        uid = s(u.get("Unit ID"))
        if uid and uid not in placed:
            unplaced_branch["children"].append(node(uid))
    if unplaced_branch["children"]:
        root["children"].append(unplaced_branch)
    return root

def compute_insights(stations_detail, day_night_station, amb_by_station, hourly_data):
    """Coverage Health + Stress Map."""
    CLINICAL = ["ARF1","ARF2","ARF3","MUZ1","MUZ2","MUZ3","MIN1","MIN2","MIN3"]
    # Coverage Health: para-to-ambulance ratio per station
    coverage = []
    for st in CLINICAL:
        sd = stations_detail.get(st, {})
        paras = sd.get("total_size", 0)
        amb = amb_by_station.get(st, 0)
        ratio = round(paras / amb, 1) if amb else None
        # Health rating: lower ratio = better (more amb per para). Industry sweet spot ~5-8
        if ratio is None:
            health = "no-amb"
        elif ratio <= 6: health = "good"
        elif ratio <= 10: health = "fair"
        else: health = "stretched"
        coverage.append({
            "station": st, "paras": paras, "ambulances": amb,
            "ratio": ratio, "health": health,
            "day_paras": day_night_station.get(st, {}).get("day_paras", 0),
            "night_paras": day_night_station.get(st, {}).get("night_paras", 0),
        })
    # Stress Map: per station × movement, peak hourly load and capacity gap
    movements = sorted(set(h.get("mvt") for h in hourly_data.get("hours", []) if h.get("mvt")))
    stress = {}
    for st in CLINICAL:
        # Capacity = unique units assigned to this station (units_total)
        capacity = stations_detail.get(st, {}).get("total_size", 0)
        stress[st] = {"capacity": capacity, "by_mvt": {}}
        for mvt in movements:
            peak = 0
            for h in hourly_data.get("hours", []):
                if h.get("mvt") == mvt:
                    val = (h.get("stations", {}) or {}).get(st, 0)
                    if val > peak: peak = val
            gap = peak - capacity
            stress[st]["by_mvt"][mvt] = {"peak": peak, "gap": gap, "pct": round(peak/capacity*100) if capacity else 0}
    return {
        "coverage_health": coverage,
        "stress_map": stress,
        "movements": movements,
    }

def main():
    print("Hajj Ops Builder v8 (v11.8 schema)")

    # Auto-derive movement phase durations from start/end times. The hand-
    # maintained `duration_hrs` field had drifted (B1A stored 6 h, actual 5 h;
    # PRE-B stored 60 h, actual 107 h; etc.). Single source of truth: the
    # start_dh/start_hour/end_dh/end_hour fields.
    for _ph in MOVEMENT_PHASES:
        _ph["duration_hrs"] = compute_phase_duration(
            _ph["start_dh"], _ph["start_hour"], _ph["end_dh"], _ph["end_hour"]
        )

    xlsx_path = download_xlsx()
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    roles = read_sheet(wb, "Roles")
    units = read_sheet(wb, "Units", r"^(PM|DPM|ADM|MDL|CHF|DCH|LOG|OCC|DPT|TRN|SUP|Mike|Alpha|Romeo|Delta)")
    staff = read_sheet(wb, "Staff")
    shifts = read_sheet(wb, "Shifts")
    schedule = read_sheet(wb, "Schedule")
    augs = read_sheet(wb, "Augmentations", r"^[A-Z]+-\d+$")
    unit_readiness = read_sheet(wb, "Unit_Readiness")
    ambulances = read_sheet(wb, "Ambulances", r"^[EBR]\d{2}$")
    print(f"  Roles: {len(roles)} · Units: {len(units)} · Staff: {len(staff)} · Shifts: {len(shifts)} · Schedule: {len(schedule)} · Ambulances: {len(ambulances)}")

    personnel = compute_personnel(roles)
    roster = compute_roster(staff)
    org = compute_org_structure(units)
    aug = compute_augmentations(augs)
    status_counts = compute_status_counts(units)
    hourly = build_hourly(staff, schedule, shifts, units, ambulances)
    stations_detail = compute_stations_detail(units, staff, unit_readiness)
    ambulance_roster, amb_by_station, amb_by_station_type = compute_ambulance_data(ambulances, units)
    units_detail = compute_units_detail(units, staff)
    schedule_grid = compute_schedule_grid(schedule, units, shifts)
    daily_view = compute_daily_view(schedule, units, shifts)
    status_live = compute_status_live(staff, schedule)
    zone_movement = compute_zone_movement(schedule, units, shifts)
    zone_day = compute_zone_day(schedule, units)
    role_views = compute_role_views(units_detail, staff, ambulance_roster)
    accommodation_live = compute_accommodation_live(units, ambulances)
    amb_dashboards = compute_ambulance_dashboards(ambulances, units, schedule, shifts)
    day_night_station = compute_day_night_per_station(units, schedule, shifts)
    org_tree = compute_org_tree(units, staff)
    insights = compute_insights(stations_detail, day_night_station, amb_by_station, hourly)

    data = {
        "refreshed_at": datetime.now(timezone.utc).strftime("%d %b %Y · %H:%M UTC"),
        "source": "Google Drive · Mobilization_Plan.xlsx (v11.8)",
        "personnel": personnel,
        "totals": {"allocated_staff_shifts": _count_allocated_shifts(schedule, units), "movements": len(MOVEMENTS)},
        "calendar": CALENDAR,
        "timeline": TIMELINE,
        "gp_coverage": GP_COVERAGE,
        "accommodation": ACCOMMODATION,
        "org_structure": org,
        "movements": MOVEMENTS,
        "movement_phases": MOVEMENT_PHASES,
        # Metro reference (SAR official ops doc, 1447H) — used by the
        # /metro.html operational viewer to show train movements, active
        # platforms per station, expected pax flow, and tafweej zones.
        "metro": {
            "phases": MOVEMENT_PHASES,
            "platforms": METRO_PLATFORMS,
            "pax_flow": [
                {"movement_group":g, "dh":dh, "hour":h, "station":st, "count":c}
                for g, rows in METRO_PAX_FLOW.items()
                for (dh, h, st, c) in rows
            ],
            "tafweej": METRO_TAFWEEJ,
            "station_hours": METRO_STATION_HOURS,
        },
        "status_counts": status_counts,
        "zone_by_mvt": {},
        "roster": roster,
        "augmentations": aug,
        "hourly": hourly,
        "stations_detail": stations_detail,
        "ambulance_roster": ambulance_roster,
        "amb_by_station": amb_by_station,
        "amb_by_station_type": amb_by_station_type,
        "units_detail": units_detail,
        "schedule_grid": schedule_grid,
        "daily_view": daily_view,
        "status_live": status_live,
        "zone_movement": zone_movement,
        "zone_day": zone_day,
        "role_views": role_views,
        "accommodation_live": accommodation_live,
        "amb_dashboards": amb_dashboards,
        "day_night_station": day_night_station,
        "org_tree": org_tree,
        "insights": insights,
    }

    with open("data.json", "w") as f:
        # Custom default: openpyxl returns datetime.time / datetime.date /
        # datetime.datetime objects for time- and date-typed cells (e.g. the
        # shift Start/End columns). Stringify them rather than crashing.
        def _json_default(o):
            if hasattr(o, 'isoformat'):
                return o.isoformat()
            return str(o)
        json.dump(data, f, ensure_ascii=False, indent=None, separators=(",", ":"), default=_json_default)

    size = os.path.getsize("data.json")
    print(f"  ✓ Wrote data.json ({size:,} bytes)")
    print(f"  Roster fill: {roster['all_filled']}/{roster['all_total']} ({roster['all_fill_pct']}%)")
    print(f"  Hourly: {len(hourly['hours'])} rows · peak ARF={hourly['peak_arafat']} MUZ={hourly['peak_muzdalifah']} MIN={hourly['peak_mina']}")


    # ── /api/v1/*.json — static API endpoints ────────────────────────
    import os as _os
    _os.makedirs("api/v1", exist_ok=True)
    _api_endpoints = {
        "api/v1/health.json": {
            "status": "ok",
            "version": "v11.8",
            "build": "v8 schema",
            "service": "hajj-ops",
            "refreshed_at": data["refreshed_at"],
        },
        "api/v1/personnel.json": {
            "personnel": data["personnel"],
            "totals": data["totals"],
            "refreshed_at": data["refreshed_at"],
        },
        "api/v1/units.json": {
            "units": data.get("units_detail", []),
            "count": len(data.get("units_detail", [])),
            "refreshed_at": data["refreshed_at"],
        },
        "api/v1/stations.json": {
            "stations": data.get("stations_detail", []),
            "count": len(data.get("stations_detail", [])),
            "refreshed_at": data["refreshed_at"],
        },
        "api/v1/movements.json": {
            "movements": data.get("movements", []),
            "augmentations": data.get("augmentations", []),
            "refreshed_at": data["refreshed_at"],
        },
        "api/v1/ambulances.json": {
            "ambulances": data.get("ambulance_roster", []),
            "by_station": data.get("amb_by_station", {}),
            "count": len(data.get("ambulance_roster", [])),
            "refreshed_at": data["refreshed_at"],
        },
        "api/v1/calendar.json": {
            "calendar": data.get("calendar", []),
            "timeline": data.get("timeline", []),
            "refreshed_at": data["refreshed_at"],
        },
        "api/v1/metro.json": {
            "metro": data.get("metro", {}),
            "refreshed_at": data["refreshed_at"],
        },
        "api/v1/index.json": {
            "api": "hajj-ops v1",
            "version": "v11.8",
            "endpoints": [
                "/api/v1/health.json",
                "/api/v1/personnel.json",
                "/api/v1/units.json",
                "/api/v1/stations.json",
                "/api/v1/movements.json",
                "/api/v1/ambulances.json",
                "/api/v1/calendar.json",
                "/api/v1/metro.json",
                "/data.json",
            ],
            "docs": "https://hajj.shuki.tech/api-docs.html",
        },
    }
    for _path, _payload in _api_endpoints.items():
        with open(_path, "w") as _f:
            json.dump(_payload, _f, ensure_ascii=False, separators=(",", ":"), default=_json_default)
    print(f"  \u2713 Wrote {len(_api_endpoints)} /api/v1/*.json files")

    try: os.unlink(xlsx_path)
    except: pass

if __name__ == "__main__":
    main()
