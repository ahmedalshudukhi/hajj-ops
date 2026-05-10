#!/usr/bin/env python3
"""
Hajj CAD — Import planning data from Mobilization_Plan.xlsx → D1.

Reads the xlsx (downloaded from Drive public URL) and inserts:
  - Augmentations (raw paramedic redeployment rows, ~118)
  - Mobilization Plan (pivoted from Schedule tab: unit_id × slot_key × value)
  - Staff Assignment (roster)

Idempotent: uses INSERT OR REPLACE / DELETE+INSERT.

Usage:
  python3 migration/import_mob_sheet.py [--xlsx /path/to/file.xlsx]
"""
import os
import sys
import re
import time
import argparse
import subprocess
import urllib.request
from typing import List, Any

DEFAULT_XLSX_URL = "https://docs.google.com/uc?export=download&id=1USar5JRbsZR_YAjW_XnPSHvYLYFOYuOl"
D1_NAME = "hajj_cad"

def sql_str(s: Any) -> str:
    if s is None or s == '':
        return 'NULL'
    if isinstance(s, bool):
        return '1' if s else '0'
    if isinstance(s, (int, float)):
        if isinstance(s, float) and s == int(s):
            return str(int(s))
        return str(s)
    s = str(s).replace("'", "''").replace("\n", " ").replace("\r", "")
    return f"'{s}'"

def d1_exec_file(filepath: str) -> dict:
    cmd = ['wrangler', 'd1', 'execute', D1_NAME, '--remote', '--file', filepath]
    try:
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=180)
        out = r.stdout + r.stderr
        if 'rows_written' in out and '"success": true' in out:
            return {'ok': True, 'output': out}
        return {'ok': False, 'output': out}
    except Exception as e:
        return {'ok': False, 'error': str(e)}

def parse_slot_key(slot_key: str) -> tuple:
    """Parse '4DH-S1' → (4, 1). Returns (None, None) on failure."""
    m = re.match(r'(\d+)DH-S(\d+)', slot_key)
    if m:
        return int(m.group(1)), int(m.group(2))
    return None, None

# ============ Augmentations ============
def import_augmentations(wb) -> int:
    print("\n[1/3] Importing Augmentations...")
    ws = wb['Augmentations']
    headers = [c.value for c in ws[1]]
    # Find column indexes
    col_idx = {h: i for i, h in enumerate(headers) if h}
    required = ['Aug ID', 'From Unit', 'To Station', 'Movement', 'DH Day', 'Hour', 'Status']
    for r in required:
        if r not in col_idx:
            print(f"  ERROR: missing column '{r}'. Have: {list(col_idx.keys())[:10]}")
            return 0
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        aug_id = row[col_idx['Aug ID']]
        if not aug_id or not str(aug_id).startswith('AUG-'):
            continue
        rows.append({
            'aug_id': aug_id,
            'from_unit': row[col_idx['From Unit']],
            'to_station': row[col_idx['To Station']],
            'movement': row[col_idx['Movement']],
            'dh_day': row[col_idx['DH Day']],
            'hour': row[col_idx['Hour']],
            'status': row[col_idx['Status']] or 'Planned',
            'reason': row[col_idx.get('Reason', -1)] if 'Reason' in col_idx else None,
            'notes': row[col_idx.get('Notes', -1)] if 'Notes' in col_idx else None,
        })
    print(f"  Found {len(rows)} augmentation rows")
    if not rows:
        return 0
    sqls = ['DELETE FROM augmentations;']
    for r in rows:
        sqls.append(
            f"INSERT INTO augmentations (aug_id, from_unit, to_station, movement, dh_day, hour, status, reason, notes) VALUES ("
            f"{sql_str(r['aug_id'])}, {sql_str(r['from_unit'])}, {sql_str(r['to_station'])}, {sql_str(r['movement'])}, "
            f"{sql_str(r['dh_day'])}, {sql_str(r['hour'])}, {sql_str(r['status'])}, {sql_str(r['reason'])}, {sql_str(r['notes'])});"
        )
    return write_and_run(sqls, '/tmp/d1_aug.sql', len(rows), 'augmentations')

# ============ Mobilization Plan (pivoted) ============
def import_mobilization_plan(wb) -> int:
    print("\n[2/3] Importing Mobilization Plan (Schedule tab, pivoted long)...")
    ws = wb['Schedule']
    headers = [c.value for c in ws[1]]
    # First 4 cols are unit metadata: Unit ID, Unit Type, Size, Home
    # Remaining cols are slot keys like "4DH-S1"
    meta_keys = ['Unit ID', 'Unit Type', 'Size', 'Home']
    slot_cols = [(i, h) for i, h in enumerate(headers) if h and h not in meta_keys and re.match(r'\d+DH-S\d+', str(h))]
    print(f"  {len(slot_cols)} slot columns found")
    if not slot_cols:
        return 0
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        unit_id = row[0]
        if not unit_id:
            continue
        unit_type = row[1] if len(row) > 1 else None
        unit_size = row[2] if len(row) > 2 else None
        home = row[3] if len(row) > 3 else None
        for col_i, slot_key in slot_cols:
            if col_i >= len(row):
                continue
            value = row[col_i]
            if value is None or value == '':
                continue
            dh, shift = parse_slot_key(str(slot_key))
            rows.append({
                'unit_id': unit_id, 'unit_type': unit_type, 'unit_size': unit_size,
                'home': home, 'slot_key': slot_key, 'dh_day': dh, 'shift': shift, 'value': value
            })
    print(f"  {len(rows)} (unit×slot) entries with values")
    if not rows:
        return 0
    sqls = ['DELETE FROM mobilization_plan;']
    for r in rows:
        sqls.append(
            f"INSERT INTO mobilization_plan (unit_id, unit_type, unit_size, home, slot_key, dh_day, shift, value) VALUES ("
            f"{sql_str(r['unit_id'])}, {sql_str(r['unit_type'])}, {sql_str(r['unit_size'])}, {sql_str(r['home'])}, "
            f"{sql_str(r['slot_key'])}, {sql_str(r['dh_day'])}, {sql_str(r['shift'])}, {sql_str(r['value'])});"
        )
    return write_and_run(sqls, '/tmp/d1_mp.sql', len(rows), 'mobilization_plan')

# ============ Staff Assignment ============
def import_staff(wb) -> int:
    print("\n[3/3] Importing Staff Assignment...")
    ws = wb['Staff']
    headers = [c.value for c in ws[1]]
    col_idx = {h: i for i, h in enumerate(headers) if h}
    required = ['Staff ID', 'Name', 'Role']
    for r in required:
        if r not in col_idx:
            print(f"  ERROR: missing column '{r}'. Have: {list(col_idx.keys())[:10]}")
            return 0
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sid = row[col_idx['Staff ID']]
        if not sid:
            continue
        rows.append({
            'staff_id': sid,
            'name': row[col_idx['Name']],
            'role': row[col_idx['Role']],
            'unit': row[col_idx.get('Unit', -1)] if 'Unit' in col_idx else None,
            'slot': row[col_idx.get('Slot', -1)] if 'Slot' in col_idx else None,
            'phone': row[col_idx.get('Phone', -1)] if 'Phone' in col_idx else None,
            'email': row[col_idx.get('Email', -1)] if 'Email' in col_idx else None,
            'radio': row[col_idx.get('Radio Call Sign', -1)] if 'Radio Call Sign' in col_idx else None,
            'status': row[col_idx.get('Status', -1)] if 'Status' in col_idx else 'Vacant',
            'total_hours': row[col_idx.get('Total Hours', -1)] if 'Total Hours' in col_idx else None,
        })
    print(f"  {len(rows)} staff rows")
    if not rows:
        return 0
    sqls = ['DELETE FROM staff_assignment;']
    for r in rows:
        sqls.append(
            f"INSERT INTO staff_assignment (staff_id, name, role, unit, slot, phone, email, radio_call_sign, status, total_hours) VALUES ("
            f"{sql_str(r['staff_id'])}, {sql_str(r['name'])}, {sql_str(r['role'])}, {sql_str(r['unit'])}, {sql_str(r['slot'])}, "
            f"{sql_str(r['phone'])}, {sql_str(r['email'])}, {sql_str(r['radio'])}, {sql_str(r['status'])}, {sql_str(r['total_hours'])});"
        )
    return write_and_run(sqls, '/tmp/d1_staff.sql', len(rows), 'staff_assignment')

def write_and_run(sqls: List[str], tmpfile: str, expected_rows: int, table_name: str) -> int:
    """Batch SQL into chunks, run, return total inserted."""
    inserted = 0
    BATCH = 100
    # Run DELETE separately first, then INSERTs in batches
    delete_stmt = sqls[0]
    insert_stmts = sqls[1:]
    # Round 1: DELETE
    with open(tmpfile, 'w') as f:
        f.write(delete_stmt)
    r = d1_exec_file(tmpfile)
    if not r.get('ok'):
        print(f"  ✗ DELETE failed: {r.get('output','')[:200]}")
        # Continue anyway — table might already be empty
    for i in range(0, len(insert_stmts), BATCH):
        batch = insert_stmts[i:i+BATCH]
        with open(tmpfile, 'w') as f:
            f.write('\n'.join(batch))
        r = d1_exec_file(tmpfile)
        if r.get('ok'):
            inserted += len(batch)
            print(f"  ✓ Batch {i//BATCH + 1}: {len(batch)} rows", flush=True)
        else:
            print(f"  ✗ Batch {i//BATCH + 1} FAILED: {r.get('output','')[:300]}")
    print(f"  → {inserted}/{expected_rows} inserted into {table_name}")
    return inserted

# ============ Main ============
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--xlsx', default=None, help='Path to xlsx (default: download from Drive)')
    parser.add_argument('--skip-aug', action='store_true')
    parser.add_argument('--skip-mp', action='store_true')
    parser.add_argument('--skip-staff', action='store_true')
    args = parser.parse_args()

    print("=" * 60)
    print("Hajj CAD — Import planning data from Mob Sheet")
    print("=" * 60)

    xlsx_path = args.xlsx
    if not xlsx_path:
        xlsx_path = '/tmp/mob.xlsx'
        print(f"  Downloading from Drive public URL → {xlsx_path}")
        try:
            urllib.request.urlretrieve(DEFAULT_XLSX_URL, xlsx_path)
            size = os.path.getsize(xlsx_path)
            print(f"  Downloaded {size:,} bytes")
        except Exception as e:
            print(f"  FATAL: download failed: {e}")
            sys.exit(1)

    try:
        import openpyxl
    except ImportError:
        print("  Installing openpyxl...")
        subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl', '--quiet'], check=True)
        import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    results = {}
    if not args.skip_aug:
        results['augmentations'] = import_augmentations(wb)
    if not args.skip_mp:
        results['mobilization_plan'] = import_mobilization_plan(wb)
    if not args.skip_staff:
        results['staff_assignment'] = import_staff(wb)

    print("\n" + "=" * 60)
    print("Summary")
    print("=" * 60)
    for k, v in results.items():
        print(f"  {k:25s} {v} rows inserted")

    print("\nDone.")

if __name__ == '__main__':
    main()
